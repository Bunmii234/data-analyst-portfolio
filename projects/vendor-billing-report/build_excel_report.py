"""
Builds Vendor_Billing_Report.xlsx -- the Excel/Power Query-style
deliverable for this project. Demonstrates:
  - Raw Data sheet with realistic messiness (mixed date formats, vendor
    name casing/whitespace inconsistency) -- the "as received" state
  - Clean Data sheet -- normalized, ready for analysis
  - Vendor Summary + Category Summary -- pivot-table-style aggregations
  - Dashboard sheet -- native Excel charts (bar, pie, line) built with
    openpyxl, an interactive-dashboard-style view inside the workbook
    itself (no separate BI tool/license required)

Run after generate_data.py.
"""
import csv
import os
import random
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter

random.seed(21)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
csv_path = os.path.join(DATA_DIR, "vendor_invoices.csv")

with open(csv_path) as f:
    rows = list(csv.DictReader(f))

# ---------------------------------------------------------------------
# Build a deliberately messy "raw" version for the Raw Data sheet --
# mixed date formats and inconsistent vendor-name casing/whitespace,
# same story as the banking project but for vendor invoices.
# ---------------------------------------------------------------------
def messify_date(iso_date):
    d = datetime.strptime(iso_date, "%Y-%m-%d")
    if random.random() < 0.5:
        return d.strftime("%m/%d/%Y")
    return d.strftime("%d-%b-%Y")

def messify_vendor(name):
    if random.random() < 0.2:
        return "  " + name.upper() + "  "
    if random.random() < 0.15:
        return name.lower()
    return name

raw_rows = []
for r in rows:
    raw_rows.append({
        "Invoice ID": r["invoice_id"],
        "Vendor": messify_vendor(r["vendor_name"]),
        "Category": r["category"],
        "Invoice Date": messify_date(r["invoice_date"]),
        "Contracted Rate": r["contracted_rate"],
        "Billed Rate": r["billed_rate"],
        "Quantity": r["quantity"],
        "Billed Amount": r["billed_amount"],
        "Approval Status": r["approval_status"],
    })

wb = Workbook()

NAVY = "1F2430"
HEADER_FILL = PatternFill(start_color="1F2430", end_color="1F2430", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

def autosize(ws, ncols, min_width=10, max_width=40):
    for c in range(1, ncols + 1):
        col_letter = get_column_letter(c)
        max_len = min_width
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

# ---- Raw Data sheet ----
ws_raw = wb.active
ws_raw.title = "Raw Data"
headers = list(raw_rows[0].keys())
ws_raw.append(headers)
style_header(ws_raw, len(headers))
for r in raw_rows:
    ws_raw.append(list(r.values()))
ws_raw.freeze_panes = "A2"
autosize(ws_raw, len(headers))

# ---- Clean Data sheet ----
ws_clean = wb.create_sheet("Clean Data")
clean_headers = ["Invoice ID","Vendor","Category","Invoice Date","Unit Type","Contracted Rate",
                  "Billed Rate","Quantity","Expected Amount","Billed Amount","Variance","Flag","Note","Approval Status"]
ws_clean.append(clean_headers)
style_header(ws_clean, len(clean_headers))
for r in rows:
    ws_clean.append([
        r["invoice_id"], r["vendor_name"], r["category"], r["invoice_date"], r["unit_type"],
        float(r["contracted_rate"]), float(r["billed_rate"]), float(r["quantity"]),
        float(r["expected_amount"]), float(r["billed_amount"]), float(r["variance"]),
        r["flag"], r["note"], r["approval_status"],
    ])
ws_clean.freeze_panes = "A2"
autosize(ws_clean, len(clean_headers))
# currency format on money columns
for col in ("F","G","I","J","K"):
    for cell in ws_clean[col][1:]:
        cell.number_format = '$#,##0.00'

# ---- Vendor Summary (pivot-style) ----
vendor_agg = defaultdict(lambda: {"invoices": 0, "billed": 0.0, "expected": 0.0, "variance": 0.0, "flagged": 0})
for r in rows:
    v = r["vendor_name"]
    vendor_agg[v]["invoices"] += 1
    vendor_agg[v]["billed"] += float(r["billed_amount"])
    vendor_agg[v]["expected"] += float(r["expected_amount"])
    vendor_agg[v]["variance"] += float(r["variance"])
    if r["flag"] != "clean":
        vendor_agg[v]["flagged"] += 1

ws_vendor = wb.create_sheet("Vendor Summary")
vheaders = ["Vendor", "Invoices", "Flagged Invoices", "Flagged %", "Total Billed", "Total Expected", "Total Variance"]
ws_vendor.append(vheaders)
style_header(ws_vendor, len(vheaders))
vendor_rows_sorted = sorted(vendor_agg.items(), key=lambda kv: kv[1]["variance"], reverse=True)
for v, d in vendor_rows_sorted:
    flagged_pct = d["flagged"] / d["invoices"] if d["invoices"] else 0
    ws_vendor.append([v, d["invoices"], d["flagged"], flagged_pct, d["billed"], d["expected"], d["variance"]])
ws_vendor.freeze_panes = "A2"
for col in ("E","F","G"):
    for cell in ws_vendor[col][1:]:
        cell.number_format = '$#,##0.00'
for cell in ws_vendor["D"][1:]:
    cell.number_format = '0.0%'
autosize(ws_vendor, len(vheaders))

# ---- Category Summary (pivot-style) ----
cat_agg = defaultdict(lambda: {"invoices": 0, "billed": 0.0, "expected": 0.0, "variance": 0.0})
for r in rows:
    c = r["category"]
    cat_agg[c]["invoices"] += 1
    cat_agg[c]["billed"] += float(r["billed_amount"])
    cat_agg[c]["expected"] += float(r["expected_amount"])
    cat_agg[c]["variance"] += float(r["variance"])

ws_cat = wb.create_sheet("Category Summary")
cheaders = ["Category", "Invoices", "Total Billed", "Total Expected", "Total Variance"]
ws_cat.append(cheaders)
style_header(ws_cat, len(cheaders))
cat_rows_sorted = sorted(cat_agg.items(), key=lambda kv: kv[1]["variance"], reverse=True)
for c, d in cat_rows_sorted:
    ws_cat.append([c, d["invoices"], d["billed"], d["expected"], d["variance"]])
for col in ("C","D","E"):
    for cell in ws_cat[col][1:]:
        cell.number_format = '$#,##0.00'
autosize(ws_cat, len(cheaders))

# ---- Monthly trend (for line chart) ----
month_agg = defaultdict(lambda: {"billed": 0.0, "expected": 0.0})
for r in rows:
    m = r["invoice_date"][:7]
    month_agg[m]["billed"] += float(r["billed_amount"])
    month_agg[m]["expected"] += float(r["expected_amount"])

ws_month = wb.create_sheet("Monthly Trend")
ws_month.append(["Month", "Total Billed", "Total Expected"])
style_header(ws_month, 3)
for m in sorted(month_agg.keys()):
    ws_month.append([m, month_agg[m]["billed"], month_agg[m]["expected"]])
for col in ("B","C"):
    for cell in ws_month[col][1:]:
        cell.number_format = '$#,##0.00'
autosize(ws_month, 3)

# ---- Flag breakdown (for pie chart) ----
flag_counts = defaultdict(int)
for r in rows:
    flag_counts[r["flag"]] += 1
FLAG_LABELS = {
    "clean": "Clean", "rate_drift_small": "Rate Drift (Small)", "rate_drift_large": "Rate Drift (Large)",
    "quantity_padding": "Quantity Padding", "duplicate": "Duplicate Invoice", "late_fee_undisclosed": "Undisclosed Late Fee",
}
ws_flags = wb.create_sheet("Flag Breakdown")
ws_flags.append(["Flag Type", "Count"])
style_header(ws_flags, 2)
for flag, label in FLAG_LABELS.items():
    ws_flags.append([label, flag_counts.get(flag, 0)])
autosize(ws_flags, 2)

# ---- Dashboard sheet with native Excel charts ----
ws_dash = wb.create_sheet("Dashboard", 0)  # make it the first visible sheet
ws_dash.sheet_view.showGridLines = False
ws_dash["B2"] = "Vendor Billing Trend & Anomaly Dashboard"
ws_dash["B2"].font = Font(size=16, bold=True, color=NAVY)
ws_dash["B3"] = "Data as of Dec 2025 -- 1,462 invoices across 12 vendors, 6 categories"
ws_dash["B3"].font = Font(size=10, italic=True, color="666666")

total_billed = sum(float(r["billed_amount"]) for r in rows)
total_variance = sum(float(r["variance"]) for r in rows if float(r["variance"]) > 0)
flagged_count = sum(1 for r in rows if r["flag"] != "clean")

ws_dash["B5"] = "Total Billed"
ws_dash["B6"] = total_billed
ws_dash["B6"].number_format = '$#,##0'
ws_dash["B6"].font = Font(size=14, bold=True)
ws_dash["D5"] = "Total Overbilled"
ws_dash["D6"] = total_variance
ws_dash["D6"].number_format = '$#,##0'
ws_dash["D6"].font = Font(size=14, bold=True, color="C0392B")
ws_dash["F5"] = "Flagged Invoices"
ws_dash["F6"] = flagged_count
ws_dash["F6"].font = Font(size=14, bold=True)
ws_dash["H5"] = "Flagged Rate"
ws_dash["H6"] = flagged_count / len(rows)
ws_dash["H6"].number_format = '0.0%'
ws_dash["H6"].font = Font(size=14, bold=True)

for col, width in (("A",3),("B",16),("C",4),("D",20),("E",4),("F",14),("G",4),("H",12)):
    ws_dash.column_dimensions[col].width = width

ws_dash.print_area = "A1:U55"
ws_dash.page_setup.fitToWidth = 1
ws_dash.page_setup.fitToHeight = 0
ws_dash.sheet_properties.pageSetUpPr.fitToPage = True

# Bar chart: top 10 vendors by variance (from Vendor Summary sheet)
bar = BarChart()
bar.title = "Total Variance by Vendor (Top 10)"
bar.y_axis.title = "Variance ($)"
bar.style = 10
data_ref = Reference(ws_vendor, min_col=7, min_row=1, max_row=11)  # Total Variance col, top 10 rows
cats_ref = Reference(ws_vendor, min_col=1, min_row=2, max_row=11)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.width, bar.height = 18, 10
ws_dash.add_chart(bar, "B9")

# Pie chart: flag breakdown
pie = PieChart()
pie.title = "Invoice Flag Breakdown"
data_ref = Reference(ws_flags, min_col=2, min_row=1, max_row=7)
cats_ref = Reference(ws_flags, min_col=1, min_row=2, max_row=7)
pie.add_data(data_ref, titles_from_data=True)
pie.set_categories(cats_ref)
pie.width, pie.height = 14, 10
ws_dash.add_chart(pie, "B50")

# Line chart: monthly billed vs expected
line = LineChart()
line.title = "Monthly Billed vs. Expected"
line.y_axis.title = "$"
data_ref = Reference(ws_month, min_col=2, max_col=3, min_row=1, max_row=13)
cats_ref = Reference(ws_month, min_col=1, min_row=2, max_row=13)
line.add_data(data_ref, titles_from_data=True)
line.set_categories(cats_ref)
line.width, line.height = 18, 10
ws_dash.add_chart(line, "B30")

tmp_path = "/tmp/Vendor_Billing_Report_build.xlsx"
wb.save(tmp_path)
final_path = os.path.join(BASE_DIR, "Vendor_Billing_Report.xlsx")
import shutil
shutil.copyfile(tmp_path, final_path)
print("Workbook written.")
print(f"Total billed: ${total_billed:,.2f}")
print(f"Total overbilled (flagged): ${total_variance:,.2f}")
print(f"Flagged invoices: {flagged_count} ({100*flagged_count/len(rows):.1f}%)")
print("\nTop 5 vendors by variance:")
for v, d in vendor_rows_sorted[:5]:
    print(f"  {v}: ${d['variance']:,.2f} variance across {d['invoices']} invoices ({d['flagged']} flagged)")
